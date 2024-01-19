import os
import pyperclip
import openpyxl
import logging
import json
import datetime
from time import sleep
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from amazoncaptcha import AmazonCaptcha

with open("last_state.json", 'r') as file:
    last_state_data = json.load(file)

def setup_chrome_driver(extension_path):
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_extension(extension_path)
    driver = webdriver.Chrome(options=chrome_options)
    driver.maximize_window()
    return driver

def handle_cookies(driver):
    try:
        # Use WebDriverWait to wait for the captcha image to be present
        image_element = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, "//div[@class='a-row a-text-center']//img")))
        link = image_element.get_attribute('src')

        # Solve the captcha
        captcha = AmazonCaptcha.fromlink(link)
        captcha_value = AmazonCaptcha.solve(captcha)

        # Find the input field and enter the captcha value
        input_field = driver.find_element(By.ID, "captchacharacters")
        input_field.send_keys(captcha_value)

        # Find the submit button and click it
        button = driver.find_element(By.CLASS_NAME, "a-button-text")
        button.click()
    except TimeoutException:
        pass
    try:
        reject_button = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.ID, 'sp-cc-rejectall-link')))
        reject_button.click()
    except TimeoutException:
        pass

def get_category_folder(main_category_name):
    
    main_folder_name = main_category_name
    
    if main_category_name == 'Electrical Goods':
        main_folder_name = "Category 1"
    elif main_category_name == 'Fashion & Accessories':
        main_folder_name = "Category 2"
    elif main_category_name == 'Home & Garden':
        main_folder_name = "Category 3"
    elif main_category_name == 'Office & Business Equipment':
        main_folder_name = "Category 4"
    elif main_category_name == 'DIY':
        main_folder_name = "Category 5"
    elif main_category_name == 'Drugstore & Cosmetics':
        main_folder_name = "Category 6"
    elif main_category_name == 'Baby & Child':
        main_folder_name = "Category 7"
    elif main_category_name == 'Sport & Leisure':
        main_folder_name = "Category 8"
    elif main_category_name == 'Pet Supplies':
        main_folder_name = "Category 9"
    elif main_category_name == 'Car & Motorbike':
        main_folder_name = "Category 10"
    elif main_category_name == 'Books, Media & Entertainment':
        main_folder_name = "Category 11"
    elif main_category_name == 'Food & Beverages':
        main_folder_name = "Category 12"
    elif main_category_name == 'Other':
        main_folder_name = "Category 13"

    return main_folder_name

def get_excel_files(main_category_name):
    excel_files = []

    main_folder_name = get_category_folder(main_category_name)
    
    if os.path.exists(main_folder_name) and os.path.isdir(main_folder_name):
        excel_files = []

        # Iterate through all subdirectories and files in the given directory
        for root, dirs, files in os.walk(main_folder_name):
            # Check if the current directory contains the specified text
            if "skript 1" in root.lower():
                
                excel_files_in_folder = [file for file in files if file.lower().endswith(('.xls', '.xlsx'))]
                    
                if excel_files_in_folder:
                        # Create the full path to the Excel file
                        latest_file = max(excel_files_in_folder, key=lambda x: os.path.getctime(os.path.join(root, x)))
                        file_path = os.path.join(root, latest_file)
                        excel_files.append(file_path)

    return excel_files

def load_amazon_urls(excel_files):
    try:
        all_amazon_urls = []

        for excel_file_path in excel_files:
            wb = openpyxl.load_workbook(excel_file_path)
            sheet = wb.active
            amazon_urls = [tuple(sheet.cell(row=cell.row, column=col).value for col in range(1, sheet.max_column + 1)) for cell in sheet['A'] if cell.value and cell.row > 1]
            all_amazon_urls.extend(amazon_urls)
            
        logging.info(f"Amazon URLs loaded from: {excel_file_path}")
        all_amazon_urls_sorted = sorted(all_amazon_urls, key=lambda x: x[2])
        return all_amazon_urls_sorted
    
    except Exception as e:
        logging.error(f"Error loading Amazon URLs from {excel_file_path}: {e}")
        raise
    
def create_or_load_workbook(category):
    
    try:
        
        main_folder_name = get_category_folder(category)
        folder_path = os.path.join(main_folder_name, f'6. Skript 2')
        os.makedirs(folder_path, exist_ok=True)

        excel_file_path = os.path.join(folder_path, f'./Script2_Germany_{category}_{datetime.datetime.now().strftime("%Y%m%d")}.xlsx')

        if os.path.isfile(excel_file_path):
            workbook = openpyxl.load_workbook(excel_file_path)
            logging.info(f"Workbook loaded: {excel_file_path}")
            worksheet = workbook.active
            last_row = worksheet.max_row + 1
        else:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.cell(1, 1).value = 'ASINs'
            worksheet.cell(1, 2).value = 'Category Name'
            worksheet.cell(1, 3).value = 'Category Level'
            worksheet.cell(1, 4).value = 'Date'
            workbook.save(excel_file_path)
            logging.info(f"New workbook created: {excel_file_path}")
            last_row = 2

        return workbook, excel_file_path, last_row

    except Exception as e:
        logging.error(f"Error creating or loading workbook: {e}")
        raise

def extract_asins_and_category(driver, url, worksheet, last_row, Category_Level):
    driver.get(url)
    sleep(3)
    handle_cookies(driver)

    try:
        asin_extractor = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.ID, 'azASINExtractorDropDown')))
        asin_extractor.click()

        all_asins_option = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, "//a[contains(@class, 'header') and contains(text(), 'All ASINs')]")))
        all_asins_option.click()

        sleep(3)
        captured_asins = pyperclip.paste()
        asins_list = captured_asins.split("\n")

        if not asins_list:
            raise ValueError("ASINs list is empty.")

        category_name = driver.find_element(By.XPATH, "//h1[contains(@class, 'a-size-large')]").text
        category_name = category_name.replace("Best Sellers in ", "")

        for value in asins_list:
            worksheet.cell(row=last_row, column=1).value = value
            worksheet.cell(row=last_row, column=2).value = category_name
            worksheet.cell(row=last_row, column=3).value = Category_Level
            worksheet.cell(row=last_row, column=4).value = datetime.datetime.now().strftime("%Y-%m-%d")
            last_row += 1
        logging.info(f"ASINs extracted from: {category_name}")
        return last_row
        # Break out of the loop if extraction is successful

    except Exception as e:
        logging.warning(f"ASINs list is empty")
        return last_row

def remove_duplicate_asins(worksheet):
    unique_asins = set()
    rows_to_delete = []

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=1):
        for cell in row:
            if cell.value in unique_asins or cell.value == None:
                rows_to_delete.append(cell.row)
                logging.warning(f"Duplicate ASIN found in row {cell.row}. Removing.")
            else:
                unique_asins.add(cell.value)

    for row_number in reversed(rows_to_delete):
        worksheet.delete_rows(row_number)

def save_workbook(workbook, excel_file_path):
    try:
        workbook.save(excel_file_path)
        logging.info(f"Workbook saved: {excel_file_path}")
    except Exception as e:
        logging.error(f"Error saving workbook: {e}")
        raise

def main(driver, category):

    try:
        excel_files = get_excel_files(category)
        workbook, asin_excel_file_path, last_row = create_or_load_workbook(category)
                
        amazon_urls = load_amazon_urls(excel_files)

        for url in amazon_urls:
            try:
                Category_URL, Category_Level = url[1], url[2]

                worksheet = workbook.active

                last_row = extract_asins_and_category(driver, Category_URL, worksheet, last_row, Category_Level)

                save_workbook(workbook, asin_excel_file_path)
            
            except Exception as e:
                logging.error(f"Error processing URL {Category_URL}: {e}")
        
        remove_duplicate_asins(worksheet)
        save_workbook(workbook, asin_excel_file_path)
        workbook.close()
        logging.info(f"Workbook closed: {asin_excel_file_path}")
        
        logging.info("Script execution completed successfully.")
    
    except Exception as e:
        logging.error(f"An unexpected error occurred: {e}")

if __name__ == "__main__":
    logging.basicConfig(filename=f'Log Files\\logfile_{datetime.datetime.now().strftime("%Y%m%d%H%M%S")}.txt ', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    extension_path = './amazoncrxextension.crx'
    driver = setup_chrome_driver(extension_path)
    handle_cookies(driver)
    main(driver, "Electrical Goods")
