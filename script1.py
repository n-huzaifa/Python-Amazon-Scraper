import os
import openpyxl
import logging
import json
import datetime
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from amazoncaptcha import AmazonCaptcha
from openpyxl.utils import get_column_letter

with open("last_state.json", 'r') as file:
    last_state_data = json.load(file)

with open("data.json", 'r') as file:
    data = json.load(file)

all_urls = data["all_urls"]

def setup_chrome_driver(extension_path):
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_extension(extension_path)
    driver = webdriver.Chrome(options=chrome_options)
    driver.maximize_window()
    return driver

def handle_cookies(driver):
    try:
        # Use WebDriverWait to wait for the captcha image to be present
        image_element = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, "//div[@class='a-row a-text-center']//img")))
        link = image_element.get_attribute('src')

        # Solve the captcha
        captcha = AmazonCaptcha.fromlink(link)
        captcha_value = AmazonCaptcha.solve(captcha)

        # Find the input field and enter the captcha value
        input_field = driver.find_element(By.ID, "captchacharacters")
        input_field.send_keys(captcha_value)

        # Find the submit button and click it
        button = driver.find_element(By.CLASS_NAME, "a-button-text")
        button.click()
    except TimeoutException:
        pass
    try:
        reject_button = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.ID, 'sp-cc-rejectall-link')))
        reject_button.click()
    except TimeoutException:
        pass

def get_category_folder(main_category_name):
    
    main_folder_name = main_category_name
    
    if main_category_name == 'Electrical Goods':
        main_folder_name = "Category 1"
    elif main_category_name == 'Fashion & Accessories':
        main_folder_name = "Category 2"
    elif main_category_name == 'Home & Garden':
        main_folder_name = "Category 3"
    elif main_category_name == 'Office & Business Equipment':
        main_folder_name = "Category 4"
    elif main_category_name == 'DIY':
        main_folder_name = "Category 5"
    elif main_category_name == 'Drugstore & Cosmetics':
        main_folder_name = "Category 6"
    elif main_category_name == 'Baby & Child':
        main_folder_name = "Category 7"
    elif main_category_name == 'Sport & Leisure':
        main_folder_name = "Category 8"
    elif main_category_name == 'Pet Supplies':
        main_folder_name = "Category 9"
    elif main_category_name == 'Car & Motorbike':
        main_folder_name = "Category 10"
    elif main_category_name == 'Books, Media & Entertainment':
        main_folder_name = "Category 11"
    elif main_category_name == 'Food & Beverages':
        main_folder_name = "Category 12"
    elif main_category_name == 'Other':
        main_folder_name = "Category 13"

    return main_folder_name

def load_or_create_workbook(main_category_name, category_type, category_type_no, excel_file_name):

    main_folder_name = get_category_folder(main_category_name)

    excel_file_path = os.path.join(main_folder_name, f'{category_type_no}. Skript 1 {category_type}', excel_file_name)
    print(excel_file_path)
    try:
        workbook = openpyxl.load_workbook(excel_file_path)
        worksheet = workbook.active
        logging.info(f"Workbook loaded successfully: {excel_file_path}")
    except FileNotFoundError:
        logging.info(f"New Workbook created: {excel_file_path}")
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

    return excel_file_path, workbook, worksheet

def scrape_and_write(driver, worksheet, level, url, row, main_category_type):
    logging.info(f"Scraping and writing data for level {level} from URL: {url}")
    
    current_level = level
    level_urls = scrape_categories(driver, url)
    for level_url in level_urls:
        worksheet.cell(row=row, column=1).value = level_url['name']
        worksheet.cell(row=row, column=2).value = level_url['url']
        worksheet.cell(row=row, column=3).value = str(current_level + 1)
        worksheet.cell(row=row, column=4).value = main_category_type
        
        row += 1
        if current_level + 1 < 3:
            row = scrape_and_write(driver, worksheet, current_level + 1, level_url['url'], row, main_category_type)

    logging.info(f"Scraping and writing data for level {level} completed.")
    return row

def scrape_categories(driver, url):
    base_url = 'https://www.amazon.de/'
    max_retries = 3
    array_to_append = []

    for _ in range(max_retries):
        try:
            
            driver.get(url)
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.CLASS_NAME, '_p13n-zg-nav-tree-all_style_zg-browse-group__88fbz')))

            soup = BeautifulSoup(driver.page_source, 'html.parser')
            group_divs = soup.find_all('div', {'role': 'group', 'class': '_p13n-zg-nav-tree-all_style_zg-browse-group__88fbz'})

            for group_div in group_divs:
                categories = []
                item_divs = group_div.find_all('div', {'role': 'treeitem', 'class': '_p13n-zg-nav-tree-all_style_zg-browse-item__1rdKf'})

                for item_div in item_divs:
                    anchor_tag = item_div.find('a')
                    span_tag = item_div.find('span')

                    if span_tag:
                        categories = []
                        break

                    if anchor_tag:
                        category_name = anchor_tag.text.strip()
                        category_url = base_url + anchor_tag['href']
                        categories.append({'name': category_name, 'url': category_url})

                for category in categories:
                    array_to_append.append({'name': category['name'], 'url': category['url']})
            break

        except TimeoutException:
            logging.warning("Timed out waiting for page to load, retrying...")
    
    return array_to_append

def remove_duplicate_categories(worksheet):
    unique_categories = set()

    for row in range(worksheet.max_row, 1, -1):
        category_name = worksheet.cell(row=row, column=1).value
        if category_name in unique_categories or category_name is None:
            logging.info(f"Removing duplicate category: {category_name}")
            worksheet.delete_rows(row)
        else:
            unique_categories.add(category_name)

    logging.info("Duplicate categories removed successfully.")

def save_workbook(workbook, worksheet, excel_file_path):
    try:
        workbook.save(excel_file_path)
        worksheet.sort(column=get_column_letter(3))
        workbook.save(excel_file_path)
        workbook.close()
        logging.info(f"Workbook saved successfully: {excel_file_path}")
    except Exception as e:
        logging.error(f"An error occurred while saving workbook {excel_file_path}: {str(e)}")

def main(driver, all_urls):
    
    try:
        for url_data in all_urls:
            url = url_data[0]
            level = url_data[1]
            # 1 - Electrical Goods - Camera & Photo - Hot New Releases
            category_data = url_data[2].split("-")

            driver.get(url)
            handle_cookies(driver)

            main_category_name = category_data[1].strip()
            category_name = category_data[2].strip()
            category_type = category_data[3].strip()
            category_type_no = category_data[0].strip()

            excel_file_name = f'./Script1_Germany_{main_category_name}_{datetime.datetime.now().strftime("%Y%m%d")}.xlsx'

            excel_file_path, workbook, worksheet = load_or_create_workbook(main_category_name, category_type, category_type_no, excel_file_name)

            worksheet.cell(row=1, column=1).value = "Category Name"
            worksheet.cell(row=1, column=2).value = "Category URL"
            worksheet.cell(row=1, column=3).value = "Category Level"
            worksheet.cell(row=1, column=4).value = "Category Type"
            worksheet.cell(row=2, column=1).value = category_name
            worksheet.cell(row=2, column=2).value = url
            worksheet.cell(row=2, column=3).value = "100" if level == 0 else str(level)
            worksheet.cell(row=2, column=4).value = category_type
            workbook.save(excel_file_path)

            row = 3
            row = scrape_and_write(driver, worksheet, level, url, row, category_type)

            remove_duplicate_categories(worksheet)
            save_workbook(workbook, worksheet, excel_file_path)

    except Exception as e:
        logging.error(f"An error occurred while scraping URL {url}: {str(e)}")

if __name__ == "__main__":

    logging.basicConfig(filename=f'Log Files\\logfile_{datetime.datetime.now().strftime("%Y%m%d%H%M%S")}.txt ', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    base_url = 'https://www.amazon.de/'
    extension_path = './amazoncrxextension.crx'
    driver = setup_chrome_driver(extension_path)
    handle_cookies(driver)
    
    main(driver, all_urls)
