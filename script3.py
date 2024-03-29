import smtplib
import os
import openpyxl
import re
import logging
import datetime
import json
from time import sleep
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from amazoncaptcha import AmazonCaptcha

def setup_chrome_driver():
    chrome_options = webdriver.ChromeOptions()
    extension_path = './amazoncrxextension.crx'
    chrome_options.add_extension(extension_path)
    driver = webdriver.Chrome(options=chrome_options)
    driver.maximize_window()
    return driver

def handle_cookies(driver):
    try:
        # Use WebDriverWait to wait for the captcha image to be present
        image_element = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, "//div[@class='a-row a-text-center']//img")))
        link = image_element.get_attribute('src')

        # Solve the captcha
        captcha = AmazonCaptcha.fromlink(link)
        captcha_value = AmazonCaptcha.solve(captcha)

        # Find the input field and enter the captcha value
        input_field = driver.find_element(By.ID, "captchacharacters")
        input_field.send_keys(captcha_value)

        # Find the submit button and click it
        button = driver.find_element(By.CLASS_NAME, "a-button-text")
        button.click()
    except TimeoutException:
        pass
    try:
        reject_button = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.ID, 'sp-cc-rejectall-link')))
        reject_button.click()
    except TimeoutException:
        pass

def get_last_state():
    try:
        global last_state_data
        with open("last_state.json", 'r') as file:
            last_state_data = json.load(file)
        return last_state_data['last_category'], last_state_data['last_asin']
    except:
        return "Electrical Goods", None

def get_category_folder(category):
    category_mapping = {
        'Electrical Goods': "Category 1",
        'Fashion & Accessories': "Category 2",
        'Home & Garden': "Category 3",
        'Office & Business Equipment': "Category 4",
        'DIY': "Category 5",
        'Drugstore & Cosmetics': "Category 6",
        'Baby & Child': "Category 7",
        'Sport & Leisure': "Category 8",
        'Pet Supplies': "Category 9",
        'Car & Motorbike': "Category 10",
        'Books, Media & Entertainment': "Category 11",
        'Food & Beverages': "Category 12",
        'Other': "Category 13"
    }
    return category_mapping.get(category, None)

def get_excel_file_path(main_category_name):

    main_folder_name = get_category_folder(main_category_name)

    folder_path = os.path.join(main_folder_name, f'6. Skript 2')
    
    excel_files = []
    if os.path.exists(folder_path) and os.path.isdir(folder_path):
    # Iterate through the files in the directory
        for file in os.listdir(folder_path):
            # Check if the file has a .xls or .xlsx extension
            if file.lower().endswith(('.xls', '.xlsx')):
                # Create the full path to the Excel file
                file_path = os.path.join(folder_path, file)
                excel_files.append(file_path)
    
    # Return the latest Excel file based on modification time
    if excel_files:
        latest_file = max(excel_files, key=lambda x: os.path.getctime(x))
        return latest_file
    else:
        return None  # Return None if no Excel files were found

def load_excel_workbook(last_asin, excel_file_path):
    wb = openpyxl.load_workbook(excel_file_path)
    sheet = wb.active
    asins = [tuple(sheet.cell(row=cell.row, column=col).value for col in range(1, sheet.max_column + 1)) for cell in sheet['A'] if cell.value and cell.row > 1]
    if last_asin is not None:
        try:
            # Find the index of the last ASIN in the list
            last_asin_index = next((index for index, a in enumerate(asins) if a[0] == last_asin), None)
            
            if last_asin_index is not None:
                # Return ASINs that come after the last ASIN
                return asins[last_asin_index + 1:]
        except:
            pass

    return asins

def create_or_load_workbook(category):
    
    main_folder_name = get_category_folder(category)
    folder_path = os.path.join(main_folder_name, f'7. Skript 3')
    os.makedirs(folder_path, exist_ok=True)
    excel_file_path = os.path.join(folder_path, f'./Script3_Germany_{category}_{datetime.datetime.now().strftime("%Y%m%d")}.xlsx')
    
    if os.path.isfile(excel_file_path):
        workbook = openpyxl.load_workbook(excel_file_path)
        worksheet = workbook.active
        row_index = worksheet.max_row
    else:
        workbook = openpyxl.Workbook()
        workbook.save(excel_file_path)
        worksheet = workbook.active

        worksheet.cell(row=1, column=1).value = "ASIN"
        worksheet.cell(row=1, column=2).value = "Brand"
        worksheet.cell(row=1, column=3).value = "Title"
        worksheet.cell(row=1, column=4).value = "Price"
        worksheet.cell(row=1, column=5).value = "Best Seller Rank"
        worksheet.cell(row=1, column=6).value = "Review Numbers"
        worksheet.cell(row=1, column=7).value = "Sales Figure"
        worksheet.cell(row=1, column=8).value = "Avg Review"
        worksheet.cell(row=1, column=9).value = "Category Rank"
        worksheet.cell(row=1, column=10).value = 'Date'
        row_index = 2

    return workbook, worksheet, row_index, excel_file_path

def create_email(sender_email, recipient, subject, message, excel_file):
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = recipient
    msg['Subject'] = subject
    msg.attach(MIMEText(message, 'plain'))

    with open(excel_file, 'rb') as file_content:
        excel_attachment = MIMEApplication(file_content.read(), _subtype="xlsx")
    excel_attachment.add_header('content-disposition', 'attachment', filename=excel_file)
    msg.attach(excel_attachment)
    logging.info(f"file attached: {excel_file}")

    return msg

def send_mail(sender_email, sender_password, recipient, subject, message, excel_files):
    smtp_server = smtplib.SMTP('smtp.gmail.com', 587)
    smtp_server.starttls()

    try:
        smtp_server.login(sender_email, sender_password)
        email_msg = create_email(sender_email, recipient, subject, message, excel_files)
        smtp_server.sendmail(sender_email, recipient, email_msg.as_string())
        logging.info("Email sent successfully.")
    except Exception as e:
        logging.error(f"Failed to send email. Error: {e}")
    finally:
        smtp_server.quit()

def main_mail(excel_file):
    with open("data.json", 'r') as file:
        data = json.load(file)

    recipient_email = data["recipient_email"]
    sender_email = data["sender_email"]
    sender_password = data["sender_password"]
    subject = 'Excel File'
    message = 'Please find the attached Excel file.'

    send_mail(sender_email, sender_password, recipient_email, subject, message, excel_file)

def extract_price(driver):
    # Try to find the price directly on the page
    try:
        product_price_element = driver.find_element(By.ID, "price_inside_buybox")
        return product_price_element.text.strip()
    except:
        pass

    # Get the page source after clicking "See All Buying Options"
    page_source = driver.page_source

    # Parse the page source with BeautifulSoup
    soup = BeautifulSoup(page_source, 'html.parser')

    # Find the elements that contain the price
    price_whole = soup.find('span', class_='a-price-whole')
    price_fraction = soup.find('span', class_='a-price-fraction')

    # Extract the price values
    if price_whole and price_fraction:
        try:
            euro_symbol = soup.find('span', class_='a-price-symbol')
            return euro_symbol.text + price_whole.text + price_fraction.text
        except:
            pass

    # If price still not found, look for "Temporarily out of stock" message
    out_of_stock_element = soup.find('span', class_='a-color-price.a-text-bold')
    try:
        out_of_stock = out_of_stock_element.text.strip()
        if "Temporarily out of stock" in out_of_stock:
            return "Temporarily out of stock"
    except:
        pass

    return "Price Not Found"

def extract_product_data(category, driver, asin, row_index, worksheet):
    
    try:
        title, brand, price, best_seller_rank, reviewNumbers, sales_figure, avgReview, rank_category = (None, None, None, None, None, None, None, None)

        page_html = driver.page_source
        soup = BeautifulSoup(page_html, 'html.parser')
        
        # Find the Sales Figures
        sales_figures_element = soup.find("span", {"class": "dpWidgetSoldUnitsLabelSum"})
        if sales_figures_element:
            try:
                # Extract the Sales Figures value
                sales_figure = sales_figures_element.find("span").text.strip()
            except:
                sales_figure = None  # Set to None if data is not found
        
        # Find the Best Seller Rank
        best_seller_rank_element = soup.find("span", class_="dpWidgetBestSellerRankLabelSum")
        if best_seller_rank_element:
            try:
                # Extract the Best Sellers Rank text
                rank_text = best_seller_rank_element.text.strip()

                # Use regex to extract the numeric part from the rank_text
                match = re.search(r'(\d+(?:,\d+)*)', rank_text)

                if match:
                    best_seller_rank = match.group(1)
                else:
                    best_seller_rank = 'Best Sellers Rank not found'
            except:
                best_seller_rank = 'Best Sellers Rank not found'

        if sales_figure and best_seller_rank != 'Best Sellers Rank not found':
            # Find the Title
            product_title_element = soup.find("span", {"id": "productTitle"})
            if product_title_element:
                try:
                    title = product_title_element.text.strip()
                except:
                    title = None  # Set to None if data is not found

            # Find the Brand
            product_brand_element = soup.find("a", {"id": "bylineInfo"})
            if product_brand_element:
                try:
                    full_brand = product_brand_element.text.strip()

                    # Extract the brand name as you described
                    brand_parts = full_brand.split()

                    if len(brand_parts) > 2:
                    # Omit the first two words and the last word
                        brand = ' '.join(brand_parts[2:-1])
                    else:
                        brand = full_brand  # Use the original brand if it doesn't match the expected format

                except Exception as e:
                    # Log or handle the exception appropriately
                    print(f"Error extracting brand: {e}")
                    brand = None

            # Find the Review Number
            product_number_of_reviews_element = soup.find("span", {"id": "acrCustomerReviewText"})
            if product_number_of_reviews_element:
                try:
                    reviewNumbers = product_number_of_reviews_element.text.strip()
                except:
                    reviewNumbers = None  # Set to None if data is not found
        
            # Find the Average Review
            product_avg_review_element = soup.find("div", {"id": "averageCustomerReviews"})
            if product_number_of_reviews_element:
                try:
                    avg_rating_sentence = product_avg_review_element.find("span", {"class": "a-declarative"}).text.strip()
                    avgReview = avg_rating_sentence[:3]
                except:
                    avgReview = None  # Set to None if data is not found

            # Find the Rank Category
            rank_category_elemet = soup.find('div', id='wayfinding-breadcrumbs_feature_div')
            if rank_category_elemet:
                try:                
                    rank_category = rank_category_elemet.find('a').text.strip()
                except:
                    rank_category = None  # Set to None if data is not found
        
            # Extract price
            price = extract_price(driver)

            # Extract data and update the Excel sheet
            row_index = update_excel_sheet(category, worksheet, row_index, asin, brand, title, price, best_seller_rank, reviewNumbers, sales_figure, avgReview, rank_category)
            logging.info(f"Successfully extracted data for ASIN: {asin}")
        else:
            logging.info(f"Sales Figure & Best Seller Rank for ASIN {asin} not found. Skipping to the next ASIN.")
            
    except Exception as e:
        logging.error(f"ASIN {asin} not found or encountered an error: {e}. Skipping to the next ASIN.")
    
    finally:
        return row_index 

def update_excel_sheet(category, worksheet, row_index, asin, brand, title, price, best_seller_rank, review_numbers, sales_figure, avg_review, category_rank):
    
    try:
        worksheet.cell(row=row_index, column=1).value = asin
        worksheet.cell(row=row_index, column=2).value = brand
        worksheet.cell(row=row_index, column=3).value = title
        worksheet.cell(row=row_index, column=4).value = price
        worksheet.cell(row=row_index, column=5).value = best_seller_rank
        worksheet.cell(row=row_index, column=6).value = review_numbers
        worksheet.cell(row=row_index, column=7).value = sales_figure
        worksheet.cell(row=row_index, column=8).value = avg_review
        worksheet.cell(row=row_index, column=9).value = category_rank
        worksheet.cell(row=row_index, column=10).value = datetime.datetime.now().strftime("%Y-%m-%d")
        row_index = row_index + 1
        last_state_data['last_category'] = category
        last_state_data['last_asin'] = asin
        with open("last_state.json", 'w') as file:
            json.dump(last_state_data, file, indent=4)
        logging.info(f"Updated Excel sheet for ASIN: {asin}")
    except Exception as e:
        logging.error(f"Error updating Excel sheet for ASIN {asin}: {e}")
    finally:
        return row_index

def main(driver):
    
    category, last_asin = get_last_state()
    try:
        excel_file_path = get_excel_file_path(category)

        if excel_file_path:
            try:
                asins = load_excel_workbook(last_asin, excel_file_path)
                
                workbook, worksheet, row_index, excel_file_path_asin = create_or_load_workbook(category)

                for asin in asins:

                    try:
                        driver.get(f'https://www.amazon.de/dp/{asin[0]}')
                        sleep(3)
                        handle_cookies(driver)
                        
                        # Use WebDriverWait with a timeout of 60 seconds for each ASIN extraction
                        WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, "productTitle")))
                        
                        row_index = extract_product_data(category, driver, asin[0], row_index, worksheet)
                        workbook.save(excel_file_path_asin)

                    except TimeoutException:
                        logging.warning(f"Timeout occurred for ASIN {asin[0]}. Skipping to the next row.")
                        continue
                    
                logging.info("Script execution completed successfully")
                
            except Exception as e:
                logging.error(f"Error processing Excel file {excel_file_path}: {e}")
        else:
                logging.error("None File with Asins there.")

        workbook.save(excel_file_path_asin)
        workbook.close()
        main_mail(excel_file_path_asin)
    except Exception as e:
        logging.error(f"Script execution failed: {e}")

if __name__ == "__main__":
    driver = setup_chrome_driver()
    handle_cookies(driver)
    main(driver, "Electrical Goods")
